import openpyxl
import os
import tkinter as tk
from tkinter import messagebox

def save_anime_titles():
    # File name for the Excel file
    file_name = "Anime_Titles.xlsx"
    
    # Create a workbook and sheet if the file doesn't exist, otherwise open the existing file
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        print("Loaded existing Excel file.")
        
        # Retrieve existing titles from the sheet
        existing_titles = []
        for row in sheet.iter_rows(values_only=True):
            if row[0] is not None:
                existing_titles.append(row[0])
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Anime Titles"
        existing_titles = []
        print("Created a new Excel file.")

    # List to hold anime titles entered by the user
    anime_titles = anime_titles_entry.get("1.0", tk.END).strip().splitlines()
    
    # Merge existing titles with the new ones and remove duplicates
    unique_anime_titles = list(set(existing_titles + anime_titles))

    # Sort titles alphabetically (optional, for better organization)
    unique_anime_titles.sort()

    # Clear existing data and write the unique anime titles to the Excel sheet
    sheet.delete_rows(1, sheet.max_row)
    for idx, title in enumerate(unique_anime_titles, start=1):
        sheet.cell(row=idx, column=1, value=title)

    # Save the workbook
    workbook.save(file_name)
    messagebox.showinfo("Success", f"Anime titles saved to '{file_name}' without duplicates.")

# Set up the GUI window
root = tk.Tk()
root.title("Anime Titles Excel Saver")

# Instruction label
instructions_label = tk.Label(root, text="Enter anime titles (one per line):")
instructions_label.pack(pady=5)

# Text box for entering anime titles
anime_titles_entry = tk.Text(root, width=40, height=15)
anime_titles_entry.pack(pady=5)

# Save button
save_button = tk.Button(root, text="Save Titles to Excel", command=save_anime_titles)
save_button.pack(pady=10)

# Run the GUI loop
root.mainloop()
