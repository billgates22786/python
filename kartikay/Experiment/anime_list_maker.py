import openpyxl
import os

def create_excel_with_anime_titles():
    # File name for the Excel file
    file_name = "Anime_Titles.xlsx"
    
    # Create a workbook and sheet if the file doesn't exist, otherwise open the existing file
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        print("Loaded existing Excel file.")
        
        # Retrieve existing titles from the sheet
        existing_titles = []
        for row in sheet.iter_rows(values_only=True):
            if row[0] is not None:
                existing_titles.append(row[0])
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Anime Titles"
        existing_titles = []
        print("Created a new Excel file.")

    # List to hold anime titles entered by the user
    anime_titles = []

    # Input loop to get anime titles from the user
    print("Enter anime titles (type 'done' to finish):")
    while True:
        title = input("Enter anime title: ")
        if title.lower() == 'done':
            break
        anime_titles.append(title)

    # Merge existing titles with the new ones and remove duplicates
    unique_anime_titles = list(set(existing_titles + anime_titles))

    # Sort titles alphabetically (optional, for better organization)
    unique_anime_titles.sort()

    # Clear existing data and write the unique anime titles to the Excel sheet
    sheet.delete_rows(1, sheet.max_row)
    for idx, title in enumerate(unique_anime_titles, start=1):
        sheet.cell(row=idx, column=1, value=title)

    # Save the workbook
    workbook.save(file_name)
    print(f"Anime titles saved to '{file_name}' without duplicates.")

# Run the function
create_excel_with_anime_titles()
